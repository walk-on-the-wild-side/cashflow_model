"""
WAL calculation, Excel output, validation, and decrement tables.
"""

import re
import pandas as pd
import numpy as np
from typing import Dict, Any, Optional, List
import warnings
from datetime import datetime

# WAL Calculation 
# Verify Calculation
def compute_wal(cashflows_df: pd.DataFrame, settlement_date=None) -> float:
    """
    Compute WAL in years. WAL = Σ(t * Principal) / Σ(Principal)
    t = period / 12
    """
    if cashflows_df is None or cashflows_df.empty:
        return 0.0
    df = cashflows_df[cashflows_df["principal_payment"] > 0.01].copy()
    if df.empty:
        return 0.0
    total = df["principal_payment"].sum()
    if total < 1.0:
        return 0.0
    df["time_years"] = df["period"] / 12.0
    return round((df["time_years"] * df["principal_payment"]).sum() / total, 2)


def compute_all_wals(all_results: dict, settlement_date=None) -> pd.DataFrame:
    rows = []
    for scenario_id, class_results in all_results.items():
        for class_id, df in class_results.items():
            wal = compute_wal(df, settlement_date)
            rows.append({"scenario_id": scenario_id, "class_id": class_id, "wal_years": wal})
    return pd.DataFrame(rows)


# WAL Validation
def validate_wals(model_wals, scenario_meta, benchmark_wals, tolerance=0.15):
    rows = []
    for _, row in model_wals.iterrows():
        sid = row["scenario_id"]
        cid = row["class_id"]
        model_wal = row["wal_years"]
        psa = scenario_meta.get(sid, {}).get("psa_speed", None)
        if psa is None:
            continue
        bench_wal = benchmark_wals.get(cid, {}).get(psa, None)
        diff = round(model_wal - bench_wal, 3) if bench_wal is not None else None
        passed = abs(diff) <= tolerance if diff is not None else None
        rows.append({
            "scenario_id": sid,
            "psa_speed": psa,
            "class_id": cid,
            "model_wal": model_wal,
            "benchmark_wal": bench_wal,
            "difference": diff,
            "within_tolerance": passed,
        })
    return pd.DataFrame(rows)


# Excel Output — full workbook
def write_excel_output(all_results, collateral_results, wal_df, validation_df,
                       scenario_meta, output_path, deal_name="Deal Model Output",
                       deal_config=None,summary_sids=None):
    """
    Write comprehensive Excel workbook.

    Sheets:
      1. Summary Dashboard   — deal overview and WAL comparison
      2. WAL Validation      — model vs. prospectus benchmark tie-out
      3. WAL Pivot           — classes * scenarios pivot
      4. Collateral CF       — one tab per group per scenario
      5. Class CF            — one tab per scenario (all classes)
      6. Decrement Tables    — % original balance outstanding
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        wb = writer.book


        # Color palette 
        NAVY   = "2E4A7A"
        LBLUE  = "D0E4F7"
        ALT    = "F2F6FC"
        WHITE  = "FFFFFF"
        GREEN  = "C6EFCE"
        RED    = "FFC7CE"
        YELLOW = "FFEB9C"

        # Header Style
        def _hdr_style(ws, row=1):
            from openpyxl.styles import PatternFill, Font, Alignment
            fill = PatternFill("solid", fgColor=NAVY)
            font = Font(bold=True, color=WHITE, size=10)
            for cell in ws[row]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Fill Solid 
        def _alt_rows(ws, start=2):
            from openpyxl.styles import PatternFill
            fill = PatternFill("solid", fgColor=ALT)
            for i, row in enumerate(ws.iter_rows(min_row=start), start=start):
                if i % 2 == 0:
                    for c in row:
                        c.fill = fill

        # Set Width
        def _autowidth(ws, max_w=30):
            from openpyxl.utils import get_column_letter
            for col in ws.columns:
                w = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, max_w)
        
        # Format Numbers
        def _fmt_numbers(ws, start_row=2):
            from openpyxl.styles import numbers as opnums
            for row in ws.iter_rows(min_row=start_row):
                for cell in row:
                    if isinstance(cell.value, float):
                        if abs(cell.value) > 1000:
                            cell.number_format = '#,##0'
                        elif abs(cell.value) < 1.5:
                            cell.number_format = '0.0000'
                        else:
                            cell.number_format = '#,##0.00'

        ################### SHEET 1: SUMMARY DASHBOARD

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        ws_sum = wb.create_sheet("Summary Dashboard")

        def wc(ws, r, c, val, bold=False, bg=None, color="000000", align="left", size=11, num_fmt=None):
            from openpyxl.styles import PatternFill, Font, Alignment
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(bold=bold, color=color, size=size)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            if num_fmt:
                cell.number_format = num_fmt
            return cell

        # Title
        ws_sum.merge_cells("A1:H1")
        wc(ws_sum, 1, 1, deal_name, bold=True, bg=NAVY, color=WHITE, align="center", size=14)
        ws_sum.row_dimensions[1].height = 30

        ws_sum.merge_cells("A2:H2")
        wc(ws_sum, 2, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                        f"Cash Flow Model", bg=LBLUE, align="center", size=10)

        # Deal metadata
        if deal_config:
            settle = deal_config["deal"].get("settlement_date", "")
            total_bal = deal_config["deal"].get("total_balance", 0)
            n_groups = len(deal_config.get("groups", []))
            n_classes = len(deal_config.get("classes", []))
        else:
            settle, total_bal, n_groups, n_classes = "", 0, 0, 0

        r = 4
        wc(ws_sum, r, 1, "DEAL PARAMETERS", bold=True, bg=LBLUE, size=11)
        r += 1
        for label, val in [
            ("Settlement Date", settle),
            ("Total Deal Size", f"${total_bal:,.0f}" if total_bal else ""),
            ("Collateral Groups", n_groups),
            ("Tranche Classes", n_classes),
            ("Scenarios Run", len(scenario_meta)),
        ]:
            wc(ws_sum, r, 1, label, bold=True, bg=ALT)
            wc(ws_sum, r, 2, val)
            r += 1

        # WAL Summary Table
        r += 1
        wc(ws_sum, r, 1, "WEIGHTED AVERAGE LIFE SUMMARY (years)", bold=True, bg=LBLUE, size=11)
        r += 1

        # Build pivot
        if not wal_df.empty:
            wal_copy = wal_df[wal_df["scenario_id"].isin(summary_sids)].copy() if summary_sids else wal_df.copy()
            # wal_copy = wal_df.copy()
            wal_copy["label"] = wal_copy["scenario_id"].map(
                lambda s: scenario_meta.get(s, {}).get("label", s))
            labels = sorted(wal_copy["label"].unique(),
                            key=lambda x: wal_copy.loc[wal_copy["label"]==x, "psa_speed"].iloc[0]
                            if "psa_speed" in wal_copy.columns else 0)
            classes = sorted(wal_copy["class_id"].unique())

            # Header row
            wc(ws_sum, r, 1, "Class", bold=True, bg=NAVY, color=WHITE, align="center")
            wc(ws_sum, r, 2, "Group", bold=True, bg=NAVY, color=WHITE, align="center")
            for ci, lab in enumerate(labels, start=3):
                wc(ws_sum, r, ci, lab, bold=True, bg=NAVY, color=WHITE, align="center", size=9)
            r += 1

            # Data rows + benchmark comparison
            bench_data = deal_config.get("benchmark_wals", {}) if deal_config else {}
            for ri, cid in enumerate(classes):
                bg = ALT if ri % 2 == 0 else WHITE
                grp = next((c["group_id"] for c in (deal_config or {}).get("classes",[])
                            if c["class_id"] == cid), "")
                wc(ws_sum, r, 1, cid, bold=True, bg=bg)
                wc(ws_sum, r, 2, grp, bg=bg, align="center")
                for ci, lab in enumerate(labels, start=3):
                    sub = wal_copy[(wal_copy["class_id"]==cid) & (wal_copy["label"]==lab)]
                    val = sub["wal_years"].iloc[0] if not sub.empty else ""
                    psa = sub["psa_speed"].iloc[0] if not sub.empty else None
                    bench = bench_data.get(cid, {}).get(psa, None) if psa is not None else None
                    cell = wc(ws_sum, r, ci, val, bg=bg, align="center", num_fmt="0.00")
                    # Color code vs benchmark
                    if bench is not None and isinstance(val, (int, float)):
                        diff = abs(val - bench)
                        if diff <= 0.15:
                            cell.fill = PatternFill("solid", fgColor=GREEN)
                        elif diff <= 0.50:
                            cell.fill = PatternFill("solid", fgColor=YELLOW)
                        else:
                            cell.fill = PatternFill("solid", fgColor=RED)
                r += 1

        # Legend
        r += 1
        wc(ws_sum, r, 1, "Color Key:", bold=True)
        wc(ws_sum, r, 2, "Green = within ±0.15yr of benchmark", bg=GREEN)
        r += 1
        wc(ws_sum, r, 2, "Yellow = within ±0.50yr of benchmark", bg=YELLOW)
        r += 1
        wc(ws_sum, r, 2, "Red = more than ±0.50yr from benchmark", bg=RED)

        # Column widths
        ws_sum.column_dimensions["A"].width = 12
        ws_sum.column_dimensions["B"].width = 8
        for i in range(3, 3 + len(scenario_meta)):
            from openpyxl.utils import get_column_letter
            ws_sum.column_dimensions[get_column_letter(i)].width = 18


        ############### SHEET 2: WAL VALIDATION

        if not validation_df.empty:
            vdf = validation_df.sort_values(["class_id", "psa_speed"]).copy()
            # Add color indicator
            vdf["status"] = vdf["within_tolerance"].map(
                {True: "PASS", False: "FAIL", None: "N/A"})
            vdf.to_excel(writer, sheet_name="WAL Validation", index=False)
            ws_v = wb["WAL Validation"]
            _hdr_style(ws_v)
            # Color pass/fail
            for row in ws_v.iter_rows(min_row=2):
                status_col = None
                for ci, cell in enumerate(row, 1):
                    if ws_v.cell(1, ci).value == "status":
                        status_col = cell
                if status_col:
                    if status_col.value == "PASS":
                        for cell in row:
                            cell.fill = PatternFill("solid", fgColor=GREEN)
                    elif status_col.value == "FAIL":
                        for cell in row:
                            cell.fill = PatternFill("solid", fgColor=RED)
            _autowidth(ws_v)
            _fmt_numbers(ws_v)


        #################### SHEET 3: WAL PIVOT

        if not wal_df.empty:
            wal_copy2 = wal_df.copy()
            wal_copy2["label"] = wal_copy2["scenario_id"].map(
                lambda s: scenario_meta.get(s, {}).get("label", s))
            pivot = wal_copy2.pivot_table(
                index="class_id", columns="label", values="wal_years", aggfunc="first")
            # Sort columns by PSA speed
            col_order = sorted(pivot.columns,
                               key=lambda x: next((scenario_meta[s]["psa_speed"]
                                                   for s in scenario_meta
                                                   if scenario_meta[s]["label"] == x), 0))
            pivot = pivot[col_order]
            pivot.to_excel(writer, sheet_name="WAL Pivot")
            ws_p = wb["WAL Pivot"]
            _hdr_style(ws_p)
            _alt_rows(ws_p)
            _autowidth(ws_p)
            _fmt_numbers(ws_p)


        ################# SHEET 4 onward: COLLATERAL CASH FLOWS (per scenario, per group)

        for sid, coll_by_group in collateral_results.items():
            label = scenario_meta.get(sid, {}).get("label", sid)
            for gid, cf_df in coll_by_group.items():
                sheet_name = ''.join(c if c not in '/?*:[]{}'.replace('{','').replace('}','') else '-' for c in f'Coll Group {gid} {label}')[:31]
                disp = cf_df[[
                    "period","bop_balance","scheduled_principal","prepayment",
                    "total_principal","ptr_interest","eop_balance","cpr","smm"
                ]].copy()
                disp.columns = [
                    "Period","BOP Balance","Sched. Principal","Prepayment",
                    "Total Principal","PTR Interest","EOP Balance","CPR","SMM"
                ]
                disp.to_excel(writer, sheet_name=sheet_name, index=False)
                ws_c = wb[sheet_name]
                _hdr_style(ws_c)
                _alt_rows(ws_c)
                _autowidth(ws_c)
                _fmt_numbers(ws_c)

        #################### SHEET: CLASS CASH FLOWS (one sheet per scenario, all classes)

        for sid, class_results in all_results.items():
            label = scenario_meta.get(sid, {}).get("label", sid)
            frames = []
            for cid, df in class_results.items():
                cols = ["class_id","group_id","period","date","bop_balance",
                        "principal_payment","interest_payment","accrual_addition",
                        "eop_balance","coupon_rate_annual","sofr"]
                sub = df[[c for c in cols if c in df.columns]].copy()
                frames.append(sub)
            if frames:
                combined = pd.concat(frames, ignore_index=True)
                combined = combined.sort_values(["class_id","period"])
                combined.columns = [c.replace("_"," ").title() for c in combined.columns]
                sheet_name = ''.join(c if c not in '/?*:[]' else '-' for c in f'ClassCF {label}')[:31]
                combined.to_excel(writer, sheet_name=sheet_name, index=False)
                ws_cf = wb[sheet_name]
                _hdr_style(ws_cf)
                _alt_rows(ws_cf)
                _autowidth(ws_cf)
                _fmt_numbers(ws_cf)


        ############### SHEET: DECREMENT TABLES (% of original balance outstanding)

        _write_decrement_tables(writer, wb, all_results, scenario_meta,
                                deal_config, NAVY, WHITE, ALT)

    print(f"Excel output written: {output_path}")


def _write_decrement_tables(writer, wb, all_results, scenario_meta,
                            deal_config, NAVY, WHITE, ALT):
    
    """Build decrement tables matching prospectus format"""
    from openpyxl.styles import PatternFill, Font, Alignment

    # Annual check dates from 2026 to 2054
    check_years = list(range(2026, 2055))
    check_dates = [f"{y}-02-25" for y in check_years]  # February of each year

    all_classes = []
    if deal_config:
        all_classes = [c["class_id"] for c in deal_config.get("classes", [])
                       if not c.get("is_notional", False)]

    # One sheet per class
    for class_id in all_classes:
        rows_data = {}
        scenario_labels = []

        for sid, label_meta in scenario_meta.items():
            label = label_meta.get("label", sid)
            scenario_labels.append(label)
            cf_df = all_results.get(sid, {}).get(class_id)
            if cf_df is None or cf_df.empty:
                for cd in check_dates:
                    rows_data.setdefault(cd, {})[label] = ""
                continue

            orig_bal = cf_df["bop_balance"].iloc[0] if not cf_df.empty else 1
            if orig_bal < 1:
                continue

            for cd in check_dates:
                subset = cf_df[cf_df["date"].astype(str) <= cd]
                if subset.empty:
                    pct = 100.0
                else:
                    eop = subset["eop_balance"].iloc[-1]
                    pct = round(eop / orig_bal * 100, 0) if orig_bal > 0 else 0.0
                rows_data.setdefault(cd, {})[label] = pct if pct > 0 else 0.0

        if not rows_data:
            continue

        # Sort scenarios by PSA speed
        scenario_labels_sorted = sorted(
            scenario_labels,
            key=lambda x: next((scenario_meta[s]["psa_speed"]
                                for s in scenario_meta
                                if scenario_meta[s]["label"] == x), 0)
        )

        sheet_name = f"Decr {class_id}"[:31]
        ws_d = wb.create_sheet(sheet_name)

        # Header
        ws_d.cell(1, 1, "Date").fill = PatternFill("solid", fgColor=NAVY)
        ws_d.cell(1, 1).font = Font(bold=True, color=WHITE, size=10)
        ws_d.cell(1, 1).alignment = Alignment(horizontal="center")
        for ci, lab in enumerate(scenario_labels_sorted, start=2):
            cell = ws_d.cell(1, ci, lab)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.alignment = Alignment(horizontal="center")

        # Also show prospectus benchmark if available 
        bench = (deal_config or {}).get("benchmark_wals", {}).get(class_id, {})

        # Data rows
        for ri, cd in enumerate(check_dates, start=2):
            row_data = rows_data.get(cd, {})
            cell = ws_d.cell(ri, 1, cd[:7])  # YYYY-MM
            cell.font = Font(bold=True, size=9)
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT)
            for ci, lab in enumerate(scenario_labels_sorted, start=2):
                val = row_data.get(lab, "")
                c2 = ws_d.cell(ri, ci, val)
                c2.alignment = Alignment(horizontal="center")
                c2.number_format = '0'
                if ri % 2 == 0:
                    c2.fill = PatternFill("solid", fgColor=ALT)

        # WAL footer
        wal_row = len(check_dates) + 3
        ws_d.cell(wal_row, 1, "WAL (yrs)").font = Font(bold=True)

        from openpyxl.utils import get_column_letter
        ws_d.column_dimensions["A"].width = 12
        for ci in range(2, 2 + len(scenario_labels_sorted)):
            ws_d.column_dimensions[get_column_letter(ci)].width = 16


def compute_wal_notional(cashflows_df, ref_cashflows_df):
    """WAL for notional IO — matches reference class WAL."""
    return compute_wal(ref_cashflows_df, None)
